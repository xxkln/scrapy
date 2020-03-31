from openpyxl import *
from openpyxl.styles import *

class MakeExcel():

    def __init__(self):
        self.dict = {
        '英雄联盟': 'https://lol.qq.com/main.shtml',
        '地下城与勇士': 'https://dnf.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        '使命召唤Online': 'https://codol.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        '疾风之刃': 'https://jf.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        '冒险岛2':'https://mxd2.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        '天涯明月刀':'https://wuxia.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        '御龙在天': 'https://yl.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        '剑灵': 'https://bns.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        '流放之路': 'https://poe.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        '轩辕传奇2':  'https://xy2.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        '斗战神':  'https://dzs.qq.com/?ADTAG=media.innerenter.gamecom.navigation',
        'QQ三国': 'https://sg.qq.com/web201706/index.shtml?ADTAG=media.innerenter.gamecom.navigation',
        }
        self.excel_template()
        self.make()

    #表格样式模板
    def excel_template(self):

        # 表格边缘线的样式，thin表示有框线，框线颜色为黑色
        self.thin = Side(border_style="thin", color="000000")

        # 设置表格上下左右都有黑色框线样式
        self.all_border = Border(top=self.thin, left=self.thin, right=self.thin, bottom=self.thin)

        # 设置普通字体样式
        self.yahei_normal = Font(name=u'微软雅黑')

        #设置标题字体样式
        self.yahei_title = Font(name=u'微软雅黑',bold=True)

        # 设置链接字体样式，设置字体为 微软雅黑，单下划线，颜色为蓝色
        self.yahei_u = Font(name=u'微软雅黑', underline='single', color='0000FF')

        # 设置背景颜色，设置充满方式为充满，颜色为黄色
        self.yellow_fill = PatternFill(fill_type='solid', fgColor='FFD700')

        # 设置字体在表格中的位置，设置字体为上下区中，字符长度超过表格宽度时自动换行
        self.alignment_wrap = Alignment(wrap_text=True, horizontal='center', vertical='center')

    #设置表格样式，根据传入的样式设置传入的单元格对象的样式
    def make_cell_style(self,cell,style):

        #普通单元格样式
        if style == 'normal':
            cell.font = self.yahei_normal
            cell.border = self.all_border
            cell.alignment = self.alignment_wrap
        #标题单元格样式
        elif style == 'title':
            cell.fill = self.yellow_fill
            cell.font = self.yahei_title
            cell.border = self.all_border
            cell.alignment = self.alignment_wrap
        #地址单元格样式
        elif style == 'addr':
            cell.font = self.yahei_u
            cell.border = self.all_border
            cell.alignment = self.alignment_wrap

    #生成表格
    def make(self):
        #创建一个工作簿，选用Sheet这张表
        workbook = Workbook()
        sheet = workbook['Sheet']

        #设置标题的名称
        cell1 = sheet['A1']
        cell2 = sheet['B1']
        cell1.value = '名称'
        cell2.value = '地址'

        #设置列框
        sheet.column_dimensions['A'].width = 20.0
        sheet.column_dimensions['B'].width = 100.0

        #设置标题样式
        self.make_cell_style(cell1, 'title')
        self.make_cell_style(cell2, 'title')

        #设置游戏和地址的样式
        cur_row = 2
        for key in self.dict.keys():

            #设置名称和地址的值
            temp_cell_A = sheet.cell(row=cur_row, column=1)
            temp_cell_B = sheet.cell(row=cur_row, column=2)
            temp_cell_A.value = key
            temp_cell_B.value = self.dict[key]
            #设置地址超链接
            temp_cell_B.hyperlink = self.dict[key]
            #设置名称和地址的样式
            self.make_cell_style(temp_cell_A, 'normal')
            self.make_cell_style(temp_cell_B, 'addr')
            cur_row = cur_row + 1

        #从第六行开始折叠
        sheet.row_dimensions.group(6, cur_row, hidden=True)
        
        workbook.save('游戏.xlsx')

MakeExcel()
